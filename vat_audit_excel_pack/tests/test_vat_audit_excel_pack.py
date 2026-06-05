import base64
import io
import unittest

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    Workbook = load_workbook = None

from odoo import Command, fields
from odoo.addons.account.tests.common import AccountTestInvoicingCommon
from odoo.tests import Form, tagged


@tagged('post_install', '-at_install')
@unittest.skipUnless(Workbook and load_workbook, 'openpyxl is required for XLSX assertions')
class TestVatAuditExcelPack(AccountTestInvoicingCommon):

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.date_from = fields.Date.from_string('2025-04-01')
        cls.date_to = fields.Date.from_string('2025-06-30')
        cls.sale_tax = cls.company_data['default_tax_sale']
        cls.bank_journal = cls.company_data['default_journal_bank']
        cls.bank_sale_nature_account_uae = cls._make_account(
            '41010101',
            'Revenue - UAE Clients',
        )
        cls.bank_sale_nature_account_international = cls._make_account(
            '41010102',
            'Revenue - International Clients',
        )
        cls.bank_sale_nature_account_related = cls._make_account(
            '41020101',
            'Revenue - Related Party',
        )
        cls.bank_sale_non_revenue_account = cls._make_account(
            '41030101',
            'Revenue - Other Clients',
        )

    @classmethod
    def _make_account(cls, code, name):
        return cls.env['account.account'].create({
            'name': name,
            'code': code,
            'account_type': cls.company_data['default_account_revenue'].account_type,
            'company_ids': [Command.set(cls.env.company.ids)],
        })

    @classmethod
    def _make_currency(cls, name, symbol):
        return cls.env['res.currency'].create({
            'name': name,
            'symbol': symbol,
            'rounding': 0.01,
        })

    def _create_wizard(self, **overrides):
        values = {
            'company_id': self.env.company.id,
            'date_from': self.date_from,
            'date_to': self.date_to,
            'target_move': 'posted',
            'include_general_ledger': False,
            'include_trial_balance': False,
            'include_profit_loss': False,
            'include_balance_sheet': False,
            'include_aged_receivable': False,
            'include_aged_payable': False,
        }
        values.update(overrides)
        return self.env['vat.audit.excel.pack.wizard'].create(values)

    def _create_invoice(self, move_type, lines):
        move_form = Form(
            self.env['account.move']
            .with_company(self.env.company)
            .with_context(default_move_type=move_type)
        )
        move_form.invoice_date = self.date_from
        if not move_form._get_modifier('date', 'invisible'):
            move_form.date = self.date_from
        move_form.partner_id = self.partner_a

        for amount, taxes in lines:
            with move_form.invoice_line_ids.new() as line_form:
                line_form.name = 'test line'
                line_form.price_unit = amount
                line_form.tax_ids.clear()
                for tax in taxes:
                    line_form.tax_ids.add(tax)

        move = move_form.save()
        move.action_post()
        return move

    def _create_bank_statement_line(
        self,
        *,
        amount,
        vat_category,
        payment_ref,
        counterpart_account=None,
        partner=None,
        date=None,
        foreign_currency=None,
        amount_currency=None,
    ):
        values = {
            'date': date or self.date_from,
            'journal_id': self.bank_journal.id,
            'partner_id': (partner or self.partner_a).id,
            'payment_ref': payment_ref,
            'amount': amount,
            'vat_category': vat_category,
        }
        if foreign_currency:
            values['foreign_currency_id'] = foreign_currency.id
            values['amount_currency'] = amount_currency
        statement_line = self.env['account.bank.statement.line'].create(values)
        if counterpart_account:
            _liquidity_lines, suspense_lines, _other_lines = statement_line._seek_for_lines()
            statement_line.set_account_bank_statement_line(suspense_lines.id, counterpart_account.id)
        return statement_line

    def _create_currency_rate(self, currency, rate, *, date=None):
        return self.env['res.currency.rate'].create({
            'currency_id': currency.id,
            'rate': rate,
            'name': date or self.date_from,
            'company_id': self.env.company.root_id.id,
        })

    def _reconcile_statement_line_to_invoice(self, statement_line, invoice):
        _liquidity_lines, suspense_lines, _other_lines = statement_line.with_context(
            skip_account_move_synchronization=True
        )._seek_for_lines()
        invoice_line = invoice.line_ids.filtered(
            lambda line: line.account_type in ('asset_receivable', 'liability_payable')
        )
        suspense_lines.account_id = invoice_line.account_id
        (suspense_lines + invoice_line).reconcile()

    def _get_workbook_from_wizard(self, wizard):
        return load_workbook(io.BytesIO(base64.b64decode(wizard.file_data)), data_only=False)

    def _find_row_number(self, sheet, label):
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value == label:
                    return cell.row
        self.fail(f'Could not find row label {label!r}')

    def test_bank_vat_category_field_is_available_from_vat_pack(self):
        vat_category_field = self.env['account.bank.statement.line']._fields.get('vat_category')

        self.assertTrue(vat_category_field)
        self.assertEqual(vat_category_field.type, 'selection')
        self.assertEqual(
            vat_category_field.selection,
            [
                ('standard', 'Standard Rated Sales'),
                ('zero_rated', 'Zero Rated Sales'),
                ('exempt', 'Exempt Sales'),
                ('out_of_scope', 'Out of Scope'),
            ],
        )

    def test_invoice_sales_are_still_included(self):
        invoice = self._create_invoice(
            'out_invoice',
            [
                (10.0, self.sale_tax),
                (15.0, self.sale_tax),
            ],
        )
        wizard = self._create_wizard()

        sale_sections, _purchase_rows, _not_claimable_rows = wizard._build_vat_working_rows()

        self.assertEqual(len(sale_sections['standard_rated']), 1)
        self.assertEqual(len(sale_sections['zero_rated']), 0)
        invoice_row = sale_sections['standard_rated'][0]
        self.assertAlmostEqual(invoice_row[4], float(invoice.amount_untaxed or 0.0))
        self.assertAlmostEqual(invoice_row[6], abs(float(invoice.amount_untaxed_signed or 0.0)))
        self.assertAlmostEqual(invoice_row[8], abs(float(invoice.amount_total_signed or 0.0)))
        self.assertEqual(invoice_row[9], 'Not Paid')
        self.assertEqual(invoice_row[10], '')

    def test_sales_receipts_are_included(self):
        receipt = self._create_invoice(
            'out_receipt',
            [
                (25.0, self.sale_tax),
            ],
        )
        wizard = self._create_wizard()

        sale_sections, _purchase_rows, _not_claimable_rows = wizard._build_vat_working_rows()

        self.assertEqual(len(sale_sections['standard_rated']), 1)
        receipt_row = sale_sections['standard_rated'][0]
        self.assertEqual(receipt_row[2], receipt.name or receipt.payment_reference or receipt.ref or '/')
        self.assertAlmostEqual(receipt_row[4], float(receipt.amount_untaxed or 0.0))
        self.assertAlmostEqual(receipt_row[8], abs(float(receipt.amount_total_signed or 0.0)))

    def test_direct_bank_sales_land_in_all_four_sections(self):
        self._create_bank_statement_line(
            amount=105.0,
            vat_category='standard',
            payment_ref='BANK-STANDARD',
            counterpart_account=self.bank_sale_nature_account_uae,
        )
        self._create_bank_statement_line(
            amount=50.0,
            vat_category='zero_rated',
            payment_ref='BANK-ZERO',
            counterpart_account=self.bank_sale_nature_account_international,
        )
        self._create_bank_statement_line(
            amount=40.0,
            vat_category='exempt',
            payment_ref='BANK-EXEMPT',
            counterpart_account=self.bank_sale_nature_account_related,
        )
        self._create_bank_statement_line(
            amount=30.0,
            vat_category='out_of_scope',
            payment_ref='BANK-OOS',
            counterpart_account=self.bank_sale_nature_account_uae,
        )
        wizard = self._create_wizard()

        sale_sections, _purchase_rows, _not_claimable_rows = wizard._build_vat_working_rows()

        self.assertEqual(len(sale_sections['standard_rated']), 1)
        self.assertEqual(len(sale_sections['zero_rated']), 1)
        self.assertEqual(len(sale_sections['exempt']), 1)
        self.assertEqual(len(sale_sections['out_of_scope']), 1)

        standard_row = sale_sections['standard_rated'][0]
        self.assertEqual(standard_row[1], 'BANK-STANDARD')
        self.assertEqual(standard_row[2], '')
        self.assertAlmostEqual(standard_row[6], 100.0)
        self.assertAlmostEqual(standard_row[7], 5.0)
        self.assertAlmostEqual(standard_row[8], 105.0)
        self.assertEqual(standard_row[9], 'Paid')
        self.assertEqual(standard_row[10], self.bank_journal.display_name)

        zero_row = sale_sections['zero_rated'][0]
        exempt_row = sale_sections['exempt'][0]
        out_of_scope_row = sale_sections['out_of_scope'][0]
        self.assertAlmostEqual(zero_row[7], 0.0)
        self.assertAlmostEqual(exempt_row[7], 0.0)
        self.assertAlmostEqual(out_of_scope_row[7], 0.0)

    def test_bank_line_reconciled_to_customer_invoice_is_excluded(self):
        invoice = self._create_invoice(
            'out_invoice',
            [
                (100.0, self.sale_tax),
            ],
        )
        statement_line = self._create_bank_statement_line(
            amount=float(invoice.amount_total or 0.0),
            vat_category='standard',
            payment_ref='INVOICE-RECEIPT',
        )
        self._reconcile_statement_line_to_invoice(statement_line, invoice)
        wizard = self._create_wizard()

        sale_sections, _purchase_rows, _not_claimable_rows = wizard._build_vat_working_rows()

        self.assertEqual(len(sale_sections['standard_rated']), 1)
        self.assertTrue(
            all(row[1] != 'INVOICE-RECEIPT' for row in sale_sections['standard_rated'])
        )

    def test_negative_bank_sale_reduces_sale_totals(self):
        self._create_bank_statement_line(
            amount=105.0,
            vat_category='standard',
            payment_ref='BANK-POSITIVE',
            counterpart_account=self.bank_sale_nature_account_uae,
        )
        self._create_bank_statement_line(
            amount=-52.5,
            vat_category='standard',
            payment_ref='BANK-REFUND',
            counterpart_account=self.bank_sale_nature_account_international,
        )
        wizard = self._create_wizard()

        sale_sections, _purchase_rows, _not_claimable_rows = wizard._build_vat_working_rows()

        rows = sale_sections['standard_rated']
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0][9], 'Paid')
        self.assertEqual(rows[1][9], 'Refund')
        self.assertAlmostEqual(sum(row[6] for row in rows), 50.0)
        self.assertAlmostEqual(sum(row[7] for row in rows), 2.5)
        self.assertAlmostEqual(sum(row[8] for row in rows), 52.5)

    def test_bank_sale_uses_odoo_currency_rate_when_available(self):
        foreign_currency = self._make_currency('XRT', 'XRT')
        self._create_currency_rate(foreign_currency, 1.2)
        self._create_bank_statement_line(
            amount=210.0,
            amount_currency=200.0,
            foreign_currency=foreign_currency,
            vat_category='standard',
            payment_ref='BANK-FX-ODOO',
            counterpart_account=self.bank_sale_nature_account_uae,
        )
        wizard = self._create_wizard()
        company_currency = self.env.company.currency_id

        sale_sections, _purchase_rows, _not_claimable_rows = wizard._build_vat_working_rows()

        row = sale_sections['standard_rated'][0]
        expected_gross = company_currency.round(200.0 * 1.2)
        expected_net = company_currency.round(expected_gross / 1.05)
        expected_vat = company_currency.round(expected_gross - expected_net)
        self.assertEqual(row[1], 'BANK-FX-ODOO')
        self.assertEqual(row[2], '')
        self.assertAlmostEqual(row[4], 200.0)
        self.assertAlmostEqual(row[5], 1.2)
        self.assertAlmostEqual(row[6], expected_net)
        self.assertAlmostEqual(row[7], expected_vat)
        self.assertAlmostEqual(row[8], expected_gross)
        self.assertEqual(row[9], 'Paid')

    def test_bank_sale_refund_uses_odoo_currency_rate_when_available(self):
        foreign_currency = self._make_currency('XRF', 'XRF')
        self._create_currency_rate(foreign_currency, 1.2)
        self._create_bank_statement_line(
            amount=-105.0,
            amount_currency=-100.0,
            foreign_currency=foreign_currency,
            vat_category='standard',
            payment_ref='BANK-FX-REFUND',
            counterpart_account=self.bank_sale_nature_account_international,
        )
        wizard = self._create_wizard()
        company_currency = self.env.company.currency_id

        sale_sections, _purchase_rows, _not_claimable_rows = wizard._build_vat_working_rows()

        row = sale_sections['standard_rated'][0]
        expected_gross = company_currency.round(-100.0 * 1.2)
        expected_net = company_currency.round(expected_gross / 1.05)
        expected_vat = company_currency.round(expected_gross - expected_net)
        self.assertEqual(row[1], 'BANK-FX-REFUND')
        self.assertEqual(row[2], '')
        self.assertAlmostEqual(row[4], -100.0)
        self.assertAlmostEqual(row[5], 1.2)
        self.assertAlmostEqual(row[6], expected_net)
        self.assertAlmostEqual(row[7], expected_vat)
        self.assertAlmostEqual(row[8], expected_gross)
        self.assertEqual(row[9], 'Refund')

    def test_bank_sale_falls_back_to_booked_ratio_when_no_odoo_rate_exists(self):
        foreign_currency = self._make_currency('XRB', 'XRB')
        self._create_currency_rate(
            foreign_currency,
            1.8,
            date=fields.Date.from_string('2025-04-15'),
        )
        self._create_bank_statement_line(
            amount=126.0,
            amount_currency=100.0,
            foreign_currency=foreign_currency,
            vat_category='zero_rated',
            payment_ref='BANK-FX-FALLBACK',
            counterpart_account=self.bank_sale_nature_account_related,
        )
        wizard = self._create_wizard()

        sale_sections, _purchase_rows, _not_claimable_rows = wizard._build_vat_working_rows()

        row = sale_sections['zero_rated'][0]
        self.assertEqual(row[1], 'BANK-FX-FALLBACK')
        self.assertEqual(row[2], '')
        self.assertAlmostEqual(row[4], 100.0)
        self.assertAlmostEqual(row[5], 1.26)
        self.assertAlmostEqual(row[6], 126.0)
        self.assertAlmostEqual(row[7], 0.0)
        self.assertAlmostEqual(row[8], 126.0)
        self.assertEqual(row[9], 'Paid')

    def test_company_currency_bank_sale_keeps_exchange_rate_one(self):
        self._create_bank_statement_line(
            amount=105.0,
            vat_category='standard',
            payment_ref='BANK-COMPANY-CURRENCY',
            counterpart_account=self.bank_sale_nature_account_uae,
        )
        wizard = self._create_wizard()

        sale_sections, _purchase_rows, _not_claimable_rows = wizard._build_vat_working_rows()

        row = sale_sections['standard_rated'][0]
        self.assertEqual(row[1], 'BANK-COMPANY-CURRENCY')
        self.assertEqual(row[2], '')
        self.assertAlmostEqual(row[4], 105.0)
        self.assertAlmostEqual(row[5], 1.0)
        self.assertAlmostEqual(row[6], 100.0)
        self.assertAlmostEqual(row[7], 5.0)
        self.assertAlmostEqual(row[8], 105.0)
        self.assertEqual(row[9], 'Paid')

    def test_sale_sheet_grand_totals_sum_all_sections(self):
        self._create_bank_statement_line(
            amount=105.0,
            vat_category='standard',
            payment_ref='BANK-STANDARD',
            counterpart_account=self.bank_sale_nature_account_uae,
        )
        self._create_bank_statement_line(
            amount=50.0,
            vat_category='zero_rated',
            payment_ref='BANK-ZERO',
            counterpart_account=self.bank_sale_nature_account_international,
        )
        self._create_bank_statement_line(
            amount=40.0,
            vat_category='exempt',
            payment_ref='BANK-EXEMPT',
            counterpart_account=self.bank_sale_nature_account_related,
        )
        self._create_bank_statement_line(
            amount=30.0,
            vat_category='out_of_scope',
            payment_ref='BANK-OOS',
            counterpart_account=self.bank_sale_nature_account_uae,
        )
        wizard = self._create_wizard()
        workbook = Workbook()
        workbook.remove(workbook.active)
        used_sheet_names = set()

        sale_sections, _purchase_rows, _not_claimable_rows = wizard._build_vat_working_rows()
        wizard._write_sale_sheet(
            workbook,
            used_sheet_names,
            sheet_name='Sale',
            sections=sale_sections,
        )

        sheet = workbook['Sale']
        grand_total_row = self._find_row_number(sheet, 'Grand total')
        self.assertAlmostEqual(sheet.cell(row=grand_total_row, column=7).value, 220.0)
        self.assertAlmostEqual(sheet.cell(row=grand_total_row, column=8).value, 5.0)
        self.assertAlmostEqual(sheet.cell(row=grand_total_row, column=9).value, 225.0)

    def test_direct_bank_sales_require_allowed_revenue_nature(self):
        self._create_bank_statement_line(
            amount=105.0,
            vat_category='standard',
            payment_ref='BANK-ALLOWED',
            counterpart_account=self.bank_sale_nature_account_uae,
        )
        self._create_bank_statement_line(
            amount=50.0,
            vat_category='zero_rated',
            payment_ref='BANK-BLOCKED',
            counterpart_account=self.bank_sale_non_revenue_account,
        )
        wizard = self._create_wizard()

        sale_sections, _purchase_rows, _not_claimable_rows = wizard._build_vat_working_rows()

        self.assertEqual(len(sale_sections['standard_rated']), 1)
        self.assertEqual(sale_sections['standard_rated'][0][1], 'BANK-ALLOWED')
        self.assertEqual(sale_sections['standard_rated'][0][2], '')
        self.assertEqual(len(sale_sections['zero_rated']), 0)

    def test_direct_bank_sales_with_named_revenue_natures_are_included(self):
        bank_sale_nature_account_uae_alias = self._make_account(
            '41030102',
            'Revenue UAE Client',
        )
        bank_sale_nature_account_international_alias = self._make_account(
            '41030103',
            'Revenue International Client',
        )
        bank_sale_nature_account_related_alias = self._make_account(
            '41030104',
            'Related Party Revenue',
        )

        self._create_bank_statement_line(
            amount=105.0,
            vat_category='standard',
            payment_ref='BANK-UAE-ALIAS',
            counterpart_account=bank_sale_nature_account_uae_alias,
        )
        self._create_bank_statement_line(
            amount=50.0,
            vat_category='zero_rated',
            payment_ref='BANK-INTL-ALIAS',
            counterpart_account=bank_sale_nature_account_international_alias,
        )
        self._create_bank_statement_line(
            amount=-40.0,
            vat_category='exempt',
            payment_ref='BANK-RELATED-ALIAS',
            counterpart_account=bank_sale_nature_account_related_alias,
        )
        wizard = self._create_wizard()

        sale_sections, _purchase_rows, _not_claimable_rows = wizard._build_vat_working_rows()

        self.assertEqual(len(sale_sections['standard_rated']), 1)
        self.assertEqual(len(sale_sections['zero_rated']), 1)
        self.assertEqual(len(sale_sections['exempt']), 1)
        self.assertEqual(sale_sections['standard_rated'][0][1], 'BANK-UAE-ALIAS')
        self.assertEqual(sale_sections['standard_rated'][0][2], '')
        self.assertEqual(sale_sections['zero_rated'][0][1], 'BANK-INTL-ALIAS')
        self.assertEqual(sale_sections['zero_rated'][0][2], '')
        self.assertEqual(sale_sections['exempt'][0][1], 'BANK-RELATED-ALIAS')
        self.assertEqual(sale_sections['exempt'][0][2], '')
        self.assertEqual(sale_sections['exempt'][0][9], 'Refund')

    def test_action_generate_excel_pack_handles_no_eligible_bank_lines(self):
        wizard = self._create_wizard()

        action = wizard.action_generate_excel_pack()

        self.assertEqual(action['type'], 'ir.actions.act_url')
        self.assertTrue(wizard.file_data)

        workbook = self._get_workbook_from_wizard(wizard)
        self.assertEqual(workbook.sheetnames, ['Sale', 'Purchases', 'Not Claimable'])
        sale_sheet = workbook['Sale']
        self.assertEqual(sale_sheet['A4'].value, 'Standard Rated Sales')
        self.assertEqual(sale_sheet['K5'].value, 'Bank Name')
        self.assertIsNotNone(self._find_row_number(sale_sheet, 'Zero Rated Sales'))
        self.assertIsNotNone(self._find_row_number(sale_sheet, 'Exempt Sales'))
        self.assertIsNotNone(self._find_row_number(sale_sheet, 'Out of Scope'))
