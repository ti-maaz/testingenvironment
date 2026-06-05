import base64
import datetime
import io
from copy import copy

from odoo import _, fields, models
from odoo.exceptions import ValidationError

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = load_workbook = None
    Alignment = Border = Font = PatternFill = Side = None
    get_column_letter = None


class VatAuditExcelPackWizard(models.TransientModel):
    _name = 'vat.audit.excel.pack.wizard'
    _description = 'VAT Audit Excel Pack Wizard'
    _STANDARD_VAT_RATE = 0.05
    _DIRECT_BANK_SALE_ACCOUNT_CODE_PREFIXES = (
        '41010101',
        '41010102',
        '41020101',
    )
    _DIRECT_BANK_SALE_ACCOUNT_NAMES = frozenset({
        'Revenue - UAE Clients',
        'Revenue - International Clients',
        'Revenue - Related Party',
    })
    _DIRECT_BANK_SALE_ACCOUNT_NAME_ALIASES = (
        *_DIRECT_BANK_SALE_ACCOUNT_NAMES,
        'Revenue UAE Client',
        'Revenue UAE Clients',
        'UAE Client Revenue',
        'UAE Clients Revenue',
        'Revenue International Client',
        'Revenue International Clients',
        'International Client Revenue',
        'International Clients Revenue',
        'International Revenue',
        'Revenue Related Party',
        'Related Party Revenue',
        'Revenue Realated Party',
        'Realated Party Revenue',
    )
    _DIRECT_BANK_SALE_ACCOUNT_NORMALIZED_NAMES = frozenset(
        ''.join(character for character in account_name.casefold() if character.isalnum())
        for account_name in _DIRECT_BANK_SALE_ACCOUNT_NAME_ALIASES
    )

    _CORE_SHEET_SEQUENCE = (
        'General Ledger',
        'Trial Balance',
        'Profit and Loss',
        'Balance Sheet',
        'Aged Receivable',
        'Aged Payable',
    )
    _WORKBOOK_SEQUENCE = _CORE_SHEET_SEQUENCE + ('Sale', 'Purchases', 'Not Claimable')
    _PROFIT_LOSS_XMLID_CANDIDATES = (
        'account_reports.profit_and_loss',
        'account_reports.profit_and_loss_report',
        'account_reports.profit_and_loss_report_0',
        'account_reports.profit_and_loss_0',
    )
    _BALANCE_SHEET_XMLID_CANDIDATES = (
        'account_reports.balance_sheet',
        'account_reports.balance_sheet_report',
        'account_reports.balance_sheet_report_0',
        'account_reports.balance_sheet_0',
    )
    _SALE_SECTION_SEQUENCE = (
        'standard_rated',
        'zero_rated',
        'exempt',
        'out_of_scope',
    )
    _BANK_VAT_CATEGORY_TO_SALE_SECTION = {
        'standard': 'standard_rated',
        'zero_rated': 'zero_rated',
        'exempt': 'exempt',
        'out_of_scope': 'out_of_scope',
    }

    company_id = fields.Many2one(
        'res.company',
        string='Company',
        required=True,
        default=lambda self: self.env.company,
    )
    date_from = fields.Date(
        string='Date From',
        required=True,
        default=lambda self: fields.Date.context_today(self).replace(day=1),
    )
    date_to = fields.Date(
        string='Date To',
        required=True,
        default=lambda self: fields.Date.context_today(self),
    )
    target_move = fields.Selection(
        [
            ('all', 'All Entries'),
            ('posted', 'Posted Entries'),
        ],
        string='Target Moves',
        required=True,
        default='posted',
    )
    include_general_ledger = fields.Boolean(string='Include General Ledger', default=True)
    include_trial_balance = fields.Boolean(string='Include Trial Balance', default=True)
    include_profit_loss = fields.Boolean(string='Include Profit and Loss', default=True)
    include_balance_sheet = fields.Boolean(string='Include Balance Sheet', default=True)
    include_aged_receivable = fields.Boolean(string='Include Aged Receivable', default=True)
    include_aged_payable = fields.Boolean(string='Include Aged Payable', default=True)

    file_name = fields.Char(string='File Name', readonly=True)
    file_data = fields.Binary(string='File', readonly=True)

    def action_generate_excel_pack(self):
        self.ensure_one()
        self._ensure_openpyxl_available()
        if self.date_from > self.date_to:
            raise ValidationError(_('Date From must be earlier than or equal to Date To.'))

        workbook = Workbook()
        workbook.remove(workbook.active)
        used_sheet_names = set()

        for sheet_name, report_payload in self._build_selected_core_report_payloads():
            self._write_native_sheet_copy(
                workbook,
                used_sheet_names,
                sheet_name=sheet_name,
                source_sheet=report_payload['source_sheet'],
            )

        sale_sections, purchase_rows, not_claimable_rows = self._build_vat_working_rows()
        self._write_sale_sheet(
            workbook,
            used_sheet_names,
            sheet_name='Sale',
            sections=sale_sections,
        )
        self._write_vat_working_sheet(
            workbook,
            used_sheet_names,
            sheet_name='Purchases',
            title='Purchases',
            subtitle='VAT Working',
            section_label='Purchases',
            headers=self._get_vat_working_headers(),
            rows=purchase_rows,
        )
        self._write_vat_working_sheet(
            workbook,
            used_sheet_names,
            sheet_name='Not Claimable',
            title='Not Claimable',
            subtitle='VAT Working',
            section_label='Not Claimable',
            headers=self._get_vat_working_headers(),
            rows=not_claimable_rows,
        )

        self._reorder_workbook_sheets(workbook, self._WORKBOOK_SEQUENCE)
        return self._finalize_download(workbook)

    def _ensure_openpyxl_available(self):
        if Workbook is None or load_workbook is None:
            raise ValidationError(_('The Python package openpyxl is required to generate the Excel pack.'))

    def _selected_core_sheet_specs(self):
        specs = []
        if self.include_general_ledger:
            specs.append(('General Ledger', 'general_ledger'))
        if self.include_trial_balance:
            specs.append(('Trial Balance', 'trial_balance'))
        if self.include_profit_loss:
            specs.append(('Profit and Loss', 'profit_loss'))
        if self.include_balance_sheet:
            specs.append(('Balance Sheet', 'balance_sheet'))
        if self.include_aged_receivable:
            specs.append(('Aged Receivable', 'aged_receivable'))
        if self.include_aged_payable:
            specs.append(('Aged Payable', 'aged_payable'))
        return specs

    def _build_selected_core_report_payloads(self):
        payloads = []
        for sheet_name, report_key in self._selected_core_sheet_specs():
            report = self._get_core_report(report_key)
            options = self._build_core_report_options(report_key, report)
            payloads.append((
                sheet_name,
                {
                    'source_sheet': self._export_report_to_source_sheet(report, options),
                },
            ))
        return payloads

    def _get_core_report(self, report_key):
        if report_key == 'general_ledger':
            return self._get_report_or_raise('account_reports.general_ledger_report')
        if report_key == 'trial_balance':
            return self._get_report_or_raise('account_reports.trial_balance_report')
        if report_key == 'aged_receivable':
            return self._get_report_or_raise('account_reports.aged_receivable_report')
        if report_key == 'aged_payable':
            return self._get_report_or_raise('account_reports.aged_payable_report')
        if report_key == 'profit_loss':
            return self._get_report_from_candidates(
                self._PROFIT_LOSS_XMLID_CANDIDATES,
                ('Profit and Loss', 'Statement of profit and loss'),
            )
        if report_key == 'balance_sheet':
            return self._get_report_from_candidates(
                self._BALANCE_SHEET_XMLID_CANDIDATES,
                ('Balance Sheet', 'Statement of Financial Position'),
            )
        raise ValidationError(_('Unsupported report key: %s') % report_key)

    def _get_report_or_raise(self, xmlid):
        report = self.env.ref(xmlid, raise_if_not_found=False)
        if not report:
            raise ValidationError(_('Accounting report not found: %s') % xmlid)
        return report.with_company(self.company_id).with_context(allowed_company_ids=[self.company_id.id])

    def _get_report_from_candidates(self, xmlids, names):
        for xmlid in xmlids:
            report = self.env.ref(xmlid, raise_if_not_found=False)
            if report:
                return report.with_company(self.company_id).with_context(
                    allowed_company_ids=[self.company_id.id]
                )

        report_model = self.env['account.report'].with_context(active_test=False)
        for name in names:
            report = report_model.search([('name', '=', name)], limit=1)
            if report:
                return report.with_company(self.company_id).with_context(
                    allowed_company_ids=[self.company_id.id]
                )

        for name in names:
            report = report_model.search([('name', 'ilike', name)], limit=1)
            if report:
                return report.with_company(self.company_id).with_context(
                    allowed_company_ids=[self.company_id.id]
                )

        raise ValidationError(
            _('Unable to resolve a native accounting report for: %s') % ', '.join(names)
        )

    def _build_core_report_options(self, report_key, report):
        date_mode = 'range'
        effective_date_from = self.date_from
        effective_date_to = self.date_to
        report_options = {
            'selected_variant_id': report.id,
            'date': {
                'filter': 'custom',
                'mode': date_mode,
                'date_from': fields.Date.to_string(effective_date_from) if effective_date_from else False,
                'date_to': fields.Date.to_string(effective_date_to),
            },
            'all_entries': self.target_move == 'all',
            'unfold_all': True,
            'report_id': report.id,
        }

        if report_key in ('aged_receivable', 'aged_payable', 'balance_sheet'):
            report_options['date'] = {
                'filter': 'custom',
                'mode': 'single',
                'date_from': fields.Date.to_string(self.date_to),
                'date_to': fields.Date.to_string(self.date_to),
            }

        if report_key == 'trial_balance':
            report_options['hierarchy'] = False
            report_options['show_account'] = True

        if report_key in ('aged_receivable', 'aged_payable'):
            report_options.update({
                'aging_based_on': 'base_on_maturity_date',
                'aging_interval': 30,
                'show_currency': True,
                'show_account': True,
            })

        return report.get_options(previous_options=report_options)

    def _export_report_to_source_sheet(self, report, options):
        export_data = report.export_to_xlsx(options)
        file_content = export_data.get('file_content')
        if not file_content:
            raise ValidationError(_('Native XLSX export returned no file content for %s.') % (report.name or report.display_name))

        source_workbook = load_workbook(io.BytesIO(file_content), data_only=False)
        if not source_workbook.sheetnames:
            raise ValidationError(_('Native XLSX export did not contain any sheet for %s.') % (report.name or report.display_name))
        return source_workbook[source_workbook.sheetnames[0]]

    def _write_native_sheet_copy(self, workbook, used_sheet_names, *, sheet_name, source_sheet):
        target_sheet = workbook.create_sheet(
            title=self._get_unique_sheet_name(sheet_name, used_sheet_names),
        )
        for src_row in source_sheet.iter_rows(
            min_row=1,
            max_row=source_sheet.max_row,
            min_col=1,
            max_col=source_sheet.max_column,
        ):
            for src_cell in src_row:
                if src_cell.value is None and not src_cell.has_style and not src_cell.comment and not src_cell.hyperlink:
                    continue
                dst_cell = target_sheet.cell(
                    row=src_cell.row,
                    column=src_cell.column,
                    value=src_cell.value,
                )
                if src_cell.has_style:
                    dst_cell.font = copy(src_cell.font)
                    dst_cell.fill = copy(src_cell.fill)
                    dst_cell.border = copy(src_cell.border)
                    dst_cell.alignment = copy(src_cell.alignment)
                    dst_cell.number_format = src_cell.number_format
                    dst_cell.protection = copy(src_cell.protection)
                if src_cell.comment:
                    dst_cell.comment = copy(src_cell.comment)
                if src_cell.hyperlink:
                    dst_cell._hyperlink = copy(src_cell.hyperlink)

        for merged_range in source_sheet.merged_cells.ranges:
            target_sheet.merge_cells(str(merged_range))

        for col_key, source_dimension in source_sheet.column_dimensions.items():
            target_dimension = target_sheet.column_dimensions[col_key]
            target_dimension.width = source_dimension.width
            target_dimension.hidden = source_dimension.hidden
            target_dimension.bestFit = source_dimension.bestFit
            target_dimension.outlineLevel = source_dimension.outlineLevel
            target_dimension.collapsed = source_dimension.collapsed

        for row_key, source_dimension in source_sheet.row_dimensions.items():
            target_dimension = target_sheet.row_dimensions[row_key]
            target_dimension.height = source_dimension.height
            target_dimension.hidden = source_dimension.hidden
            target_dimension.outlineLevel = source_dimension.outlineLevel
            target_dimension.collapsed = source_dimension.collapsed

        target_sheet.freeze_panes = source_sheet.freeze_panes
        if source_sheet.auto_filter and source_sheet.auto_filter.ref:
            target_sheet.auto_filter.ref = source_sheet.auto_filter.ref

    def _get_vat_working_headers(self):
        return [
            _('Date'),
            _('Description'),
            _('Inv No'),
            _('Currency'),
            _('Amount'),
            _('Exchange Rate'),
            _('Net'),
            _('VAT'),
            _('Gross'),
            _('Payment Date'),
            _('Payment Made'),
            _('Payment Due'),
        ]

    def _get_empty_sale_sections(self):
        return {section_key: [] for section_key in self._SALE_SECTION_SEQUENCE}

    def _get_sale_section_specs(self):
        return [
            ('standard_rated', _('Standard Rated Sales')),
            ('zero_rated', _('Zero Rated Sales')),
            ('exempt', _('Exempt Sales')),
            ('out_of_scope', _('Out of Scope')),
        ]

    def _build_vat_working_rows(self):
        sale_sections = self._get_empty_sale_sections()
        purchase_rows = []
        not_claimable_rows = []

        move_env = self.env['account.move'].with_context(
            allowed_company_ids=[self.company_id.id],
            active_test=False,
        )

        sales = move_env.search(
            self._get_vat_move_domain(('out_invoice', 'out_refund', 'out_receipt')),
            order='invoice_date, id',
        )
        purchases = move_env.search(
            self._get_vat_move_domain(('in_invoice', 'in_refund')),
            order='invoice_date, id',
        )

        for move in sales:
            section_key, row_values = self._build_sale_row_for_move(move)
            sale_sections[section_key].append(row_values)

        for statement_line in self._get_direct_bank_sale_lines():
            section_key, row_values = self._build_sale_row_for_bank_statement_line(statement_line)
            sale_sections[section_key].append(row_values)

        for move in purchases:
            claimable_rows, non_claimable_rows = self._split_purchase_working_rows(move)
            purchase_rows.extend(claimable_rows)
            not_claimable_rows.extend(non_claimable_rows)

        return sale_sections, purchase_rows, not_claimable_rows

    def _get_vat_move_domain(self, move_types):
        domain = [
            ('company_id', '=', self.company_id.id),
            ('move_type', 'in', list(move_types)),
        ]
        if self.target_move == 'posted':
            domain.append(('state', '=', 'posted'))
        if self.date_from:
            domain.append(('invoice_date', '>=', self.date_from))
        if self.date_to:
            domain.append(('invoice_date', '<=', self.date_to))
        return domain

    def _get_bank_sale_line_domain(self):
        domain = [
            ('company_id', '=', self.company_id.id),
            ('state', '=', 'posted'),
            ('journal_id.type', '=', 'bank'),
            ('vat_category', 'in', list(self._BANK_VAT_CATEGORY_TO_SALE_SECTION)),
        ]
        if self.date_from:
            domain.append(('date', '>=', self.date_from))
        if self.date_to:
            domain.append(('date', '<=', self.date_to))
        return domain

    def _get_direct_bank_sale_lines(self):
        statement_line_env = self.env['account.bank.statement.line'].with_context(
            allowed_company_ids=[self.company_id.id],
            active_test=False,
        )
        statement_lines = statement_line_env.search(
            self._get_bank_sale_line_domain(),
            order='date, id',
        )
        return statement_lines.filtered(self._is_direct_bank_sale_line)

    def _is_direct_bank_sale_line(self, statement_line):
        if statement_line.vat_category not in self._BANK_VAT_CATEGORY_TO_SALE_SECTION:
            return False

        if (
            abs(float(statement_line.amount or 0.0)) <= 0.00001
            and abs(float(statement_line.amount_currency or 0.0)) <= 0.00001
        ):
            return False

        if not self._statement_line_has_allowed_direct_sale_nature(statement_line):
            return False

        counterpart_moves = self._get_statement_line_counterpart_moves(statement_line)
        if counterpart_moves.filtered(lambda move: move.is_sale_document(include_receipts=True)):
            return False

        counterpart_payments = counterpart_moves.mapped('origin_payment_id')
        if counterpart_payments.filtered(
            lambda payment: payment.reconciled_invoice_ids.filtered(
                lambda move: move.is_sale_document(include_receipts=True)
            )
        ):
            return False

        return True

    def _statement_line_has_allowed_direct_sale_nature(self, statement_line):
        _liquidity_lines, _suspense_lines, other_lines = statement_line._seek_for_lines()
        return any(
            self._is_allowed_direct_bank_sale_account(account)
            for account in other_lines.account_id
        )

    def _is_allowed_direct_bank_sale_account(self, account):
        account_code = (account.code or '').strip()
        if any(
            account_code.startswith(prefix)
            for prefix in self._DIRECT_BANK_SALE_ACCOUNT_CODE_PREFIXES
        ):
            return True
        normalized_account_name = ''.join(
            character for character in (account.name or '').casefold() if character.isalnum()
        )
        return normalized_account_name in self._DIRECT_BANK_SALE_ACCOUNT_NORMALIZED_NAMES

    def _get_statement_line_counterpart_moves(self, statement_line):
        counterpart_moves = self.env['account.move']
        for line in statement_line.line_ids.filtered(lambda move_line: not move_line.display_type):
            counterpart_moves |= line.matched_debit_ids.mapped('credit_move_id.move_id')
            counterpart_moves |= line.matched_credit_ids.mapped('debit_move_id.move_id')
        return counterpart_moves - statement_line.move_id

    def _build_working_rows_for_move(self, move):
        rows = []
        payment_dates = ', '.join(self._get_move_payment_dates(move))
        exchange_rate = self._compute_move_exchange_rate(move)
        payment_made = float(move.amount_total or 0.0) - float(move.amount_residual or 0.0)
        payment_due = float(move.amount_residual or 0.0)

        for move_line in move.invoice_line_ids.sorted(lambda line: (line.sequence, line.id)):
            if move_line.display_type in {'line_section', 'line_subsection', 'line_note'}:
                continue
            vat_amount = float((move_line.price_total or 0.0) - (move_line.price_subtotal or 0.0))
            rows.append({
                'line': move_line,
                'values': [
                    move.invoice_date or move.date,
                    move.partner_id.display_name or '',
                    move.name or move.payment_reference or move.ref or '/',
                    (move.currency_id or move.company_currency_id).display_name or '',
                    float(move_line.price_subtotal or 0.0),
                    exchange_rate,
                    float(move_line.price_subtotal or 0.0) * exchange_rate,
                    vat_amount * exchange_rate,
                    float(move_line.price_total or 0.0) * exchange_rate,
                    payment_dates,
                    payment_made,
                    payment_due,
                ],
            })
        return rows

    def _build_sale_row_for_move(self, move):
        sign = -1.0 if move.move_type == 'out_refund' else 1.0
        exchange_rate = self._compute_move_exchange_rate(move)
        amount_currency = sign * float(move.amount_untaxed or 0.0)
        net_amount = sign * abs(float(move.amount_untaxed_signed or 0.0))
        gross_amount = sign * abs(float(move.amount_total_signed or 0.0))
        vat_amount = gross_amount - net_amount

        section_key = 'standard_rated' if abs(vat_amount) > 0.00001 else 'zero_rated'
        status = _('Paid') if move.payment_state in ('paid', 'in_payment') else _('Not Paid')
        row_values = [
            move.invoice_date or move.date,
            move.partner_id.display_name or '',
            move.name or move.payment_reference or move.ref or '/',
            (move.currency_id or move.company_currency_id).display_name or '',
            amount_currency,
            exchange_rate,
            net_amount,
            vat_amount,
            gross_amount,
            status,
            '',
        ]
        return section_key, row_values

    def _build_sale_row_for_bank_statement_line(self, statement_line):
        section_key = self._BANK_VAT_CATEGORY_TO_SALE_SECTION[statement_line.vat_category]
        amount_currency = self._get_bank_statement_line_transaction_amount(statement_line)
        exchange_rate, uses_odoo_rate = self._compute_bank_statement_line_exchange_rate(
            statement_line,
            amount_currency=amount_currency,
        )
        net_amount, vat_amount, gross_amount = self._compute_bank_statement_line_company_amounts(
            statement_line,
            amount_currency=amount_currency,
            exchange_rate=exchange_rate,
            uses_odoo_rate=uses_odoo_rate,
        )

        display_currency = self._get_bank_statement_line_display_currency(statement_line)
        row_values = [
            statement_line.date,
            statement_line.payment_ref or statement_line.name or '',
            '',
            display_currency.display_name or display_currency.name or '',
            amount_currency,
            exchange_rate,
            net_amount,
            vat_amount,
            gross_amount,
            _('Refund') if gross_amount < 0 else _('Paid'),
            statement_line.journal_id.display_name or statement_line.journal_id.name or '',
        ]
        return section_key, row_values

    def _split_purchase_working_rows(self, move):
        claimable = []
        non_claimable = []
        for row in self._build_working_rows_for_move(move):
            vat_value = row['values'][7] or 0.0
            if abs(vat_value) > 0.00001:
                claimable.append(row['values'])
            else:
                non_claimable.append(row['values'])
        return claimable, non_claimable

    def _compute_move_exchange_rate(self, move):
        move_currency = move.currency_id
        company_currency = move.company_currency_id
        if not move_currency or not company_currency:
            return 1.0
        if move_currency == company_currency:
            return 1.0

        currency_rate = getattr(move, 'currency_rate', 0.0) or 0.0
        if currency_rate:
            return float(currency_rate)

        total_currency = float(move.amount_total or 0.0)
        total_company = abs(float(move.amount_total_signed or 0.0))
        if total_currency:
            return total_company / total_currency
        return 1.0

    def _get_bank_statement_line_display_currency(self, statement_line):
        return (
            statement_line.foreign_currency_id
            or statement_line.currency_id
            or statement_line.company_id.currency_id
        )

    def _get_bank_statement_line_transaction_amount(self, statement_line):
        if statement_line.foreign_currency_id:
            return float(statement_line.amount_currency or 0.0)
        return float(statement_line.amount or 0.0)

    def _compute_bank_statement_line_exchange_rate(self, statement_line, *, amount_currency=None):
        display_currency = self._get_bank_statement_line_display_currency(statement_line)
        company_currency = statement_line.company_id.currency_id
        if not display_currency or display_currency == company_currency:
            return 1.0, False

        odoo_rate = self._get_bank_statement_line_odoo_exchange_rate(statement_line)
        if odoo_rate:
            return odoo_rate, True

        transaction_amount = (
            amount_currency
            if amount_currency is not None
            else self._get_bank_statement_line_transaction_amount(statement_line)
        )
        company_amount = float(statement_line.amount or 0.0)
        if transaction_amount:
            return abs(company_amount / transaction_amount), False
        return 1.0, False

    def _get_bank_statement_line_odoo_exchange_rate(self, statement_line):
        display_currency = self._get_bank_statement_line_display_currency(statement_line)
        company = statement_line.company_id
        company_currency = company.currency_id
        rate_date = statement_line.date or fields.Date.context_today(self)
        if not display_currency or display_currency == company_currency:
            return False

        rate = self.env['res.currency.rate'].search(
            [
                ('currency_id', '=', display_currency.id),
                ('name', '<=', rate_date),
                ('company_id', 'in', [False, company.root_id.id]),
            ],
            order='company_id desc, name desc, id desc',
            limit=1,
        )
        if not rate:
            return False

        conversion_rate = self.env['res.currency']._get_conversion_rate(
            display_currency,
            company_currency,
            company,
            rate_date,
        )
        return float(conversion_rate or 0.0) or False

    def _compute_bank_statement_line_company_amounts(
        self,
        statement_line,
        *,
        amount_currency,
        exchange_rate,
        uses_odoo_rate,
    ):
        company_currency = statement_line.company_id.currency_id
        gross_amount = float(statement_line.amount or 0.0)
        if uses_odoo_rate:
            gross_amount = company_currency.round(float(amount_currency or 0.0) * float(exchange_rate or 0.0))

        if statement_line.vat_category == 'standard':
            net_amount = company_currency.round(gross_amount / (1.0 + self._STANDARD_VAT_RATE)) if gross_amount else 0.0
            vat_amount = company_currency.round(gross_amount - net_amount)
        else:
            net_amount = gross_amount
            vat_amount = 0.0
        return net_amount, vat_amount, gross_amount

    def _get_move_payment_dates(self, move):
        move.invalidate_recordset(['invoice_payments_widget'])
        widget = move.invoice_payments_widget or {}
        if not isinstance(widget, dict):
            return []

        payment_dates = []
        for payment_line in widget.get('content') or []:
            raw_date = payment_line.get('date')
            if not raw_date:
                continue
            try:
                normalized = fields.Date.to_string(fields.Date.to_date(raw_date))
            except Exception:
                normalized = str(raw_date)
            if normalized not in payment_dates:
                payment_dates.append(normalized)
        return payment_dates

    def _write_vat_working_sheet(self, workbook, used_sheet_names, *, sheet_name, title, subtitle, section_label, headers, rows):
        sheet = workbook.create_sheet(title=self._get_unique_sheet_name(sheet_name, used_sheet_names))
        styles = self._get_vat_sheet_styles()

        sheet['D1'] = self.company_id.display_name or self.company_id.name or ''
        sheet['D1'].font = styles['title_font']
        sheet['D1'].alignment = styles['title_alignment']
        sheet['D2'] = subtitle
        sheet['D2'].font = styles['title_font']
        sheet['D2'].alignment = styles['title_alignment']
        sheet['D3'] = f"Period {fields.Date.to_string(self.date_from)} to {fields.Date.to_string(self.date_to)}"
        sheet['D3'].font = styles['title_font']
        sheet['D3'].alignment = styles['title_alignment']
        sheet['A4'] = section_label
        sheet['A4'].font = styles['section_font']

        header_row = 5
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=header_row, column=col_idx, value=header)
            cell.font = styles['sale_header_font']
            cell.fill = styles['sale_header_fill']
            cell.alignment = styles['header_alignment']
            cell.border = styles['border']

        write_row = header_row + 1
        for line_values in rows:
            for col_idx, value in enumerate(line_values, start=1):
                cell = sheet.cell(row=write_row, column=col_idx, value=value)
                cell.border = styles['border']
                if col_idx == 1 and isinstance(value, datetime.date):
                    if not isinstance(value, datetime.datetime):
                        value = datetime.datetime.combine(value, datetime.time.min)
                        cell.value = value
                    cell.number_format = 'yyyy-mm-dd'
                elif col_idx in (5, 6, 7, 8, 9, 11, 12) and isinstance(value, (int, float)):
                    cell.number_format = '#,##0.00'
                cell.alignment = styles['data_alignment']
            write_row += 1

        if rows:
            total_row = write_row + 1
            sheet.cell(row=total_row, column=2, value=_('Total')).font = styles['header_font']
            sheet.cell(row=total_row, column=2).border = styles['border']
            for col_idx in (5, 7, 8, 9, 11, 12):
                letter = get_column_letter(col_idx)
                total_cell = sheet.cell(
                    row=total_row,
                    column=col_idx,
                    value=f'=SUM({letter}{header_row + 1}:{letter}{write_row - 1})',
                )
                total_cell.number_format = '#,##0.00'
                total_cell.font = styles['header_font']
                total_cell.fill = styles['total_fill']
                total_cell.border = styles['border']
                total_cell.alignment = styles['data_alignment']

        width_map = {
            1: 14,
            2: 32,
            3: 16,
            4: 12,
            5: 14,
            6: 14,
            7: 14,
            8: 14,
            9: 14,
            10: 18,
            11: 14,
            12: 14,
        }
        for col_idx, width in width_map.items():
            sheet.column_dimensions[get_column_letter(col_idx)].width = width

        sheet.auto_filter.ref = f"A{header_row}:L{max(write_row - 1, header_row)}"

    def _get_sale_sheet_headers(self):
        return [
            _('Date'),
            _('Description'),
            _('Inv.No'),
            _('Currency'),
            _('Amount'),
            _('Exchange Rate'),
            _('Net'),
            _('VAT'),
            _('Gross'),
            _('Status'),
            _('Bank Name'),
        ]

    def _write_sale_sheet(self, workbook, used_sheet_names, *, sheet_name, sections):
        sheet = workbook.create_sheet(title=self._get_unique_sheet_name(sheet_name, used_sheet_names))
        styles = self._get_vat_sheet_styles()
        headers = self._get_sale_sheet_headers()

        sheet['D1'] = self.company_id.display_name or self.company_id.name or ''
        sheet['D1'].font = styles['title_font']
        sheet['D1'].alignment = styles['title_alignment']
        sheet['D2'] = _('VAT Working')
        sheet['D2'].font = styles['title_font']
        sheet['D2'].alignment = styles['title_alignment']
        sheet['D3'] = self._get_sale_period_heading()
        sheet['D3'].font = styles['title_font']
        sheet['D3'].alignment = styles['title_alignment']

        write_row = 4
        for section_key, title in self._get_sale_section_specs():
            write_row = self._write_sale_section(
                sheet,
                styles,
                start_row=write_row,
                title=title,
                headers=headers,
                rows=sections.get(section_key, []),
            )

        grand_total_row = write_row + 1
        sheet.cell(row=grand_total_row, column=2, value=_('Grand total')).font = styles['header_font']
        sheet.cell(row=grand_total_row, column=2).border = styles['border']
        grand_totals = {
            7: sum(row[6] for section_rows in sections.values() for row in section_rows),
            8: sum(row[7] for section_rows in sections.values() for row in section_rows),
            9: sum(row[8] for section_rows in sections.values() for row in section_rows),
        }
        for col_idx in (7, 8, 9):
            total_cell = sheet.cell(
                row=grand_total_row,
                column=col_idx,
                value=grand_totals[col_idx],
            )
            total_cell.number_format = '#,##0.00'
            total_cell.font = styles['header_font']
            total_cell.fill = styles['total_fill']
            total_cell.border = styles['border']
            total_cell.alignment = styles['data_alignment']

        width_map = {
            1: 14,
            2: 32,
            3: 16,
            4: 12,
            5: 14,
            6: 14,
            7: 14,
            8: 14,
            9: 14,
            10: 14,
            11: 24,
        }
        for col_idx, width in width_map.items():
            sheet.column_dimensions[get_column_letter(col_idx)].width = width

    def _write_sale_section(self, sheet, styles, *, start_row, title, headers, rows):
        section_title_cell = sheet.cell(row=start_row, column=1, value=title)
        section_title_cell.font = styles['section_font']

        header_row = start_row + 1
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=header_row, column=col_idx, value=header)
            cell.font = styles['sale_header_font']
            cell.fill = styles['sale_header_fill']
            cell.alignment = styles['header_alignment']
            cell.border = styles['border']

        write_row = header_row + 1
        for line_values in rows:
            for col_idx, value in enumerate(line_values, start=1):
                cell = sheet.cell(row=write_row, column=col_idx, value=value)
                cell.border = styles['border']
                if col_idx == 1 and isinstance(value, datetime.date):
                    if not isinstance(value, datetime.datetime):
                        value = datetime.datetime.combine(value, datetime.time.min)
                        cell.value = value
                    cell.number_format = 'yyyy-mm-dd'
                elif col_idx in (5, 6, 7, 8, 9) and isinstance(value, (int, float)):
                    cell.number_format = '#,##0.00'
                cell.alignment = styles['data_alignment']
            write_row += 1

        total_row = write_row + 1
        sheet.cell(row=total_row, column=2, value=_('Total')).font = styles['header_font']
        sheet.cell(row=total_row, column=2).border = styles['border']
        section_totals = {
            7: sum(row[6] for row in rows),
            8: sum(row[7] for row in rows),
            9: sum(row[8] for row in rows),
        }
        for col_idx in (7, 8, 9):
            total_cell = sheet.cell(
                row=total_row,
                column=col_idx,
                value=section_totals[col_idx],
            )
            total_cell.number_format = '#,##0.00'
            total_cell.font = styles['header_font']
            total_cell.fill = styles['total_fill']
            total_cell.border = styles['border']
            total_cell.alignment = styles['data_alignment']

        return total_row + 2

    def _get_sale_period_heading(self):
        if not self.date_to:
            return ''
        period_end = fields.Date.to_date(self.date_to)
        return period_end.strftime('Quarter Ended  %d %B%Y')

    def _get_vat_sheet_styles(self):
        thin_side = Side(style='thin', color='D9D9D9')
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        return {
            'title_font': Font(name='Aptos', size=12, bold=True),
            'section_font': Font(name='Aptos', size=12, bold=True, color='FF0000'),
            'header_font': Font(name='Aptos', size=11, bold=True),
            'sale_header_font': Font(name='Aptos', size=11, bold=True, color='FFFFFF'),
            'header_fill': PatternFill(fill_type='solid', fgColor='F2F2F2'),
            'sale_header_fill': PatternFill(fill_type='solid', fgColor='002060'),
            'total_fill': PatternFill(fill_type='solid', fgColor='FFF2CC'),
            'title_alignment': Alignment(horizontal='center', vertical='center'),
            'header_alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'data_alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
            'border': thin_border,
        }

    def _get_unique_sheet_name(self, base_name, used_sheet_names):
        max_len = 31
        name = (base_name or 'Sheet').strip()[:max_len]
        if name not in used_sheet_names:
            used_sheet_names.add(name)
            return name

        counter = 1
        while True:
            suffix = f" ({counter})"
            candidate = f"{name[:max_len - len(suffix)]}{suffix}"
            if candidate not in used_sheet_names:
                used_sheet_names.add(candidate)
                return candidate
            counter += 1

    def _reorder_workbook_sheets(self, workbook, ordered_sheet_names):
        ordered = [workbook[sheet_name] for sheet_name in ordered_sheet_names if sheet_name in workbook.sheetnames]
        unordered = [workbook[sheet_name] for sheet_name in workbook.sheetnames if sheet_name not in ordered_sheet_names]
        workbook._sheets = ordered + unordered

    def _finalize_download(self, workbook):
        output = io.BytesIO()
        workbook.save(output)
        file_bytes = output.getvalue()
        output.close()

        filename = (
            f"VAT_Audit_Pack_{fields.Date.to_string(self.date_from)}"
            f"_to_{fields.Date.to_string(self.date_to)}.xlsx"
        )
        self.write({
            'file_name': filename,
            'file_data': base64.b64encode(file_bytes),
        })
        return {
            'type': 'ir.actions.act_url',
            'url': (
                f"/web/content/?model={self._name}&id={self.id}"
                "&field=file_data&filename_field=file_name&download=true"
            ),
            'target': 'self',
        }
